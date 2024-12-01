from datetime import datetime, timezone
import json
from email._header_value_parser import get_domain
from math import ceil
import re
from typing import Union
from aiogram import Bot, types
import aiohttp
from aiogram.fsm.storage.base import StorageKey
from aiogram.fsm.context import FSMContext
from asgiref.sync import sync_to_async
from openpyxl import Workbook
from urllib.parse import urlparse
from modul.clientbot import strings
from modul.config import settings_conf
from modul import models

from aiogram.types import FSInputFile
import os
from modul.loader import bot_session, main_bot, dp
from aiogram import Bot


def get_fsm_context(bot: Bot, chat_id: int, user_id: int = None):
    if not user_id:
        user_id = chat_id
    return FSMContext(
        storage=dp.storage,
        key=StorageKey(
            chat_id=chat_id,
            user_id=user_id,
            bot_id=bot.id
        )
    )


def have_one_module(bot, module_name: str):
    modules = [
        "enable_promotion",
        "enable_music",
        "enable_download",
        "enable_leo",
        "enable_chatgpt",
        "enable_horoscope",
        "enable_anon",
        "enable_sms",
        "enable_refs",
        "enable_kino"
    ]
    if getattr(bot, f"enable_{module_name}"):
        return [getattr(bot, x) for x in modules].count(True) == 1
    return False


def turn_bot_data(attr: str, bot: Bot):
    bot = get_current_bot(bot)
    setattr(bot, attr, not getattr(bot, attr))
    bot.save()


@sync_to_async
def get_current_bot(bot: Bot):
    return models.Bot.objects.filter(token=bot.token).first()


@sync_to_async
def get_bot(bot: Bot):
    return models.Bot.objects.filter(token=bot.token).select_related("owner").first()


@sync_to_async
def get_bot_by_token(token: str):
    return models.Bot.objects.filter(token=token).first()


@sync_to_async
def get_bot_owner(bot: Bot):
    current_bot = get_current_bot(bot)
    return models.Bot.objects.filter(bots__token=current_bot.token).first()


@sync_to_async
def get_bot_by_username(username: str):
    return models.Bot.objects.filter(username=username).first()


@sync_to_async
def create_client_bot_user(uid, bot: Bot):
    return models.ClientBotUser.objects.filter(uid=uid, bot__token=bot.token).first()


async def get_user(uid: int, bot: Bot):
    bot = await get_current_bot(bot)
    info = await create_client_bot_user(uid, bot)
    return info


@sync_to_async
def get_users(**kwargs):
    queryset = models.ClientBotUser.objects.filter(**kwargs)
    return queryset, queryset.count()


#
#
async def increase_referral(user: models.ClientBotUser):
    user.referral_count += 1
    await user.save()


#
#
# async def referral_count(uid: int):
#     user = await get_user(uid)
#     if user:
#         return user.referral_count
#
#
# async def user_balance(uid: int):
#     user = await get_user(uid)
#     if user:
#         return user.balance
#
#
# async def referral_balance(uid: int):
#     user = await get_user(uid)
#     if user:
#         return user.referral_balance
#
#
# async def transfer_money(uid: int, amount: float):
# user = await get_user(uid)
# if user:
#     user.referral_balance -= amount
#     user.balance += amount
#     await user.save()
#     return True


# async def save_order(uid: int, order_id: int, category: str, quantity: int, price: float, profit: float,
#                      bot_admin_profit: float, link: str, service: int, category_id: int, smm: int):
#     user = await get_user(uid)
#     await models.Order.objects.create(
#         user=user,
#         order_id=order_id,
#         category=category,
#         quantity=quantity,
#         link=link,
#         price=price,
#         profit=profit,
#         status=strings.PENDING,
#         bot_admin_profit=bot_admin_profit
#     )
#     await models.OrderRetry.objects.create(
#         order_id=order_id,
#         service=service,
#         category=category_id,
#         smm=smm,
#     )

#
# async def get_orders(uid: int):
#     user = await get_user(uid)
#     if user:
#         return await models.Order.objects.filter(user=user).order_by('-id')
#     else:
#         raise Exception("User not found")


# async def update_user_balance(user_or_uid: Union[models.ClientBotUser, int], amount):
#     if isinstance(user_or_uid, int):
#         user_or_uid = await get_user(user_or_uid)
#     user_or_uid.balance += amount
#     await user_or_uid.save()


async def users_count(bot: Bot):
    # bot = await get_bot(bot)
    return await models.ClientBotUser.objects.filter(bot=bot).count()


# async def earned():
#     bot = await get_bot(Bot.get_current())
#     orders = await models.Order.objects.filter(user__bot=bot, status=strings.COMPLETED)
#     total_earned = sum(order.bot_admin_profit for order in orders)
#     return f"{total_earned:.2f}"


async def get_all_users(bot: Bot):
    bot = await get_bot(bot)
    return await models.ClientBotUser.objects.filter(bot=bot)


async def get_main_bot_user(uid: int):
    return await models.Bot.objects.filter(uid=uid).first()


async def get_all_users_count(bot: Bot):
    return await models.ClientBotUser.objects.filter(bot__token=bot.token).count()


async def get_new_users_count(bot: Bot):
    today = datetime.now(tz=timezone.utc)
    return await models.ClientBotUser.objects.filter(
        bot__token=bot.token,
        created_at__year=today.year,
        created_at__month=today.month,
        created_at__day=today.day
    ).count()


# async def add_to_favourites(user_id: int, service: int, smm: int, category: int) -> bool:
#     """Добавить в избранное"""
#     user = await get_user(user_id)
#     bot = await user.bot
#     record = await get_favourite_by_id(user_id, service)
#     if not record:
#         await models.FavouritesModel.create(
#             user=user,
#             bot=bot,
#             service=service,
#             smm=smm,
#             category=category,
#         )
#         return True
#     return False


# async def task_admin_export_history(uid: int, bot_username):
#     user = await models.MainBotUser.filter(uid=uid).first()
#     bot = await models.Bot.filter(owner=user, username=bot_username).first()
#     wb = Workbook()
#     ws_client_orders = wb.create_sheet("История")
#     ws_referals = wb.create_sheet("Рефералы")
#     ws_payouts = wb.create_sheet("Выводы")
#     ws_client_orders.append(['Клиент', 'Дата', 'Сумма', 'Заработок', ])
#     ws_referals.append(['Клиент', 'Дата', 'Сумма', 'Заработок', ])
#     ws_payouts.append(['Дата', 'Сумма', ])
#     plus = 0
#     all_payouts = 0
#     reff = 0
#     clients = await models.ClientBotUser.filter(bot=bot)
#     for client in clients:
#         orders = await models.Order.filter(status__in=["Partial", 'Completed'], user=client, )
#         for order in orders:
#             plus += order.bot_admin_profit
#             ws_client_orders.append(
#                 [client.uid, order.created_at.strftime("%d.%m.%Y"), order.price, order.bot_admin_profit])
#     ws_client_orders.append(['Итого:', '', plus, ''])
#     mes = await models.ClientBotUser.filter(uid=uid)
#     for me in mes:
#         inviters = await models.ClientBotUser.filter(inviter=me)
#         for inviter in inviters:
#             ord_inviter = await models.Order.filter(status__in=["Partial", 'Completed'], user=inviter, )
#             for order in ord_inviter:
#                 reff += order.price * 0.05
#                 ws_referals.append(
#                     [inviter.uid, order.created_at.strftime("%d.%m.%Y"), order.price, order.price * 0.05])
#     ws_referals.append(['Итого:', '', reff, ''])
#     payouts = await models.Payout.filter(user=user, payout_status="success")
#     for payout in payouts:
#         ws_payouts.append([payout.created_at.strftime("%d.%m.%Y"), payout.payout_amount])
#         all_payouts += payout.payout_amount
#     ws_payouts.append(['Итого:', all_payouts])
#     user.balance = plus - all_payouts
#     file_name = f"export-{uid}.xlsx"
#     wb.save(file_name)
#     bot_text = "Не создан"
#     if bot:
#         bot_text = f"@{bot.username}\n"
#         async with Bot(token=bot.token, session=bot_session).context(auto_close=False) as _bot:
#             await _bot.send_message(
#                 uid,
#                 f"Бот: {bot_text}\nЗаработано: {plus}\nВывод: {all_payouts}\nКлиентов: {len(inviters)}\nРефка: {reff}\nБаланс: {plus + reff - all_payouts}")
#             await _bot.send_document(uid, document=FSInputFile(file_name, filename=file_name))
#     else:
#         await main_bot.send_message(
#             uid,
#             f"Бот: Не создан\nЗаработано: {plus}\nВывод: {all_payouts}\nКлиентов: {len(inviters)}\nРефка: {reff}\nБаланс: {plus + reff - all_payouts}")
#         await main_bot.send_document(uid, document=FSInputFile(file_name, filename=file_name))
#     os.remove(file_name)


# def get_domain(url):
#     pattern = r"https?://(?:www\.)?([a-zA-Z0-9-]+)\..*"
#     match = re.match(pattern, url)
#     return match.group(1) if match else "None"

@sync_to_async
def add_to_analitic_data(bot_username: str, link: str, ignore_domain: bool = False):
    now = datetime.now()
    domain = get_domain(link) if not ignore_domain else link
    instance = models.DownloadAnalyticsModel.objects.filter(bot_username=bot_username, domain=domain,
                                                                  date__gte=now).first()
    if not instance:
        models.DownloadAnalyticsModel.objects.create(
            bot_username=bot_username,
            domain=domain,
            count=1,
            date=now,
        )
    else:
        instance.count += 1
        instance.save()
