import os
import csv
import logging
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

logger = logging.getLogger(__name__)


class AdminUtils:
    """Admin panel utility funksiyalari"""

    @staticmethod
    def export_users_to_excel(users_data: List[Dict], filename: str = None) -> str:
        """Foydalanuvchilarni Excel faylga eksport qilish"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"users_export_{timestamp}.xlsx"

            wb = Workbook()
            ws = wb.active
            ws.title = "Пользователи"

            # Header
            headers = ['ID', 'Имя', 'Баланс', 'Рефералы', 'Дата регистрации', 'Статус']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')

            # Data
            for row, user in enumerate(users_data, 2):
                ws.cell(row=row, column=1, value=user.get('id', ''))
                ws.cell(row=row, column=2, value=user.get('name', ''))
                ws.cell(row=row, column=3, value=user.get('balance', 0))
                ws.cell(row=row, column=4, value=user.get('referrals', 0))
                ws.cell(row=row, column=5, value=user.get('registration_date', ''))
                ws.cell(row=row, column=6, value=user.get('status', ''))

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            filepath = os.path.join("temp", filename)
            os.makedirs("temp", exist_ok=True)
            wb.save(filepath)

            return filepath

        except Exception as e:
            logger.error(f"Export to Excel error: {e}")
            return None

    @staticmethod
    def export_users_to_csv(users_data: List[Dict], filename: str = None) -> str:
        """Foydalanuvchilarni CSV faylga eksport qilish"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"users_export_{timestamp}.csv"

            filepath = os.path.join("temp", filename)
            os.makedirs("temp", exist_ok=True)

            with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
                fieldnames = ['id', 'name', 'balance', 'referrals', 'registration_date', 'status']
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)

                writer.writeheader()
                for user in users_data:
                    writer.writerow({
                        'id': user.get('id', ''),
                        'name': user.get('name', ''),
                        'balance': user.get('balance', 0),
                        'referrals': user.get('referrals', 0),
                        'registration_date': user.get('registration_date', ''),
                        'status': user.get('status', '')
                    })

            return filepath

        except Exception as e:
            logger.error(f"Export to CSV error: {e}")
            return None

    @staticmethod
    def format_number(number: float, decimals: int = 2) -> str:
        """Sonni formatlash"""
        try:
            if isinstance(number, (int, float)):
                return f"{number:,.{decimals}f}".replace(',', ' ')
            return str(number)
        except:
            return "0"

    @staticmethod
    def format_date(date_obj, format_str: str = "%d.%m.%Y %H:%M") -> str:
        """Sanani formatlash"""
        try:
            if hasattr(date_obj, 'strftime'):
                return date_obj.strftime(format_str)
            return str(date_obj)
        except:
            return "Не указано"

    @staticmethod
    def validate_telegram_id(tg_id: str) -> bool:
        """Telegram ID ni tekshirish"""
        try:
            id_int = int(tg_id)
            return 0 < id_int < 10 ** 10  # Telegram ID range
        except:
            return False

    @staticmethod
    def validate_amount(amount: str) -> tuple:
        """Summani tekshirish"""
        try:
            amount_float = float(amount.replace(',', '.'))
            if amount_float < 0:
                return False, "Сумма не может быть отрицательной"
            if amount_float > 1000000:
                return False, "Сумма слишком большая"
            return True, amount_float
        except ValueError:
            return False, "Некорректная сумма"

    @staticmethod
    def sanitize_text(text: str, max_length: int = 1000) -> str:
        """Matnni tozalash"""
        if not text:
            return ""

        # HTML teglarini olib tashlash
        import re
        text = re.sub(r'<[^>]+>', '', text)

        # Uzunligini cheklash
        if len(text) > max_length:
            text = text[:max_length] + "..."

        return text.strip()


# Export utils
admin_utils = AdminUtils()